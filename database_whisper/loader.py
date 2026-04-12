"""Data loading and auto-detection utilities.

Supports: CSV, TSV, JSON, NDJSON, SQLite, Excel (.xlsx), Parquet, SQL dumps.
Core formats (CSV, TSV, JSON, SQLite) need zero dependencies.
Excel needs openpyxl. Parquet needs pyarrow or pandas.
"""

from __future__ import annotations

import csv
import json
import os
import re
import sqlite3
from typing import Any, Dict, List, Optional, Tuple


def load(
    path: str,
    delimiter: Optional[str] = None,
    max_records: int = 500_000,
    table: Optional[str] = None,
    sheet: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Load any supported file into a list of dicts.

    Supported formats (auto-detected by extension):
        .csv, .tsv      — delimited text (stdlib)
        .json, .ndjson  — JSON array or newline-delimited (stdlib)
        .db, .sqlite    — SQLite database (stdlib)
        .xlsx           — Excel workbook (needs openpyxl)
        .parquet        — Apache Parquet (needs pyarrow)
        .sql            — SQL dump with INSERT statements (stdlib)

    Args:
        path: File path.
        delimiter: Override delimiter for CSV/TSV. Auto-detected if None.
        max_records: Maximum records to load.
        table: Table name for SQLite (auto-picks largest if None).
        sheet: Sheet name for Excel (first sheet if None).

    Returns:
        (records, field_names)
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".db", ".sqlite", ".sqlite3"):
        return _load_sqlite(path, max_records, table)

    if ext == ".xlsx":
        return _load_excel(path, max_records, sheet)

    if ext == ".parquet":
        return _load_parquet(path, max_records)

    if ext == ".sql":
        return _load_sql_dump(path, max_records)

    if ext in (".json", ".ndjson", ".jsonl"):
        return _load_json(path, max_records)

    # Default: CSV/TSV
    return _load_csv(path, delimiter, max_records)


# ---------------------------------------------------------------------------
# CSV / TSV (stdlib)
# ---------------------------------------------------------------------------

def _load_csv(
    path: str,
    delimiter: Optional[str],
    max_records: int,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load CSV or TSV with auto-detected delimiter."""
    if delimiter is None:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            sample = f.read(4096)
            if "\t" in sample and sample.count("\t") > sample.count(","):
                delimiter = "\t"
            else:
                delimiter = ","

    records = []
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        field_names = [fn.strip() for fn in (reader.fieldnames or []) if fn]
        for row in reader:
            if len(records) >= max_records:
                break
            record = {k.strip(): (v or "").strip() for k, v in row.items() if k}
            records.append(record)

    return records, field_names


# ---------------------------------------------------------------------------
# JSON / NDJSON (stdlib)
# ---------------------------------------------------------------------------

def _load_json(path: str, max_records: int) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load JSON array or newline-delimited JSON."""
    records = []
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read().strip()

    if content.startswith("["):
        data = json.loads(content)
        for item in data[:max_records]:
            if isinstance(item, dict):
                records.append(_flatten_dict(item))
    else:
        for line in content.split("\n"):
            if len(records) >= max_records:
                break
            line = line.strip()
            if line:
                try:
                    item = json.loads(line)
                    if isinstance(item, dict):
                        records.append(_flatten_dict(item))
                except json.JSONDecodeError:
                    continue

    field_names = list(records[0].keys()) if records else []
    return records, field_names


def _flatten_dict(d: Dict, prefix: str = "", sep: str = ".") -> Dict[str, str]:
    """Flatten a nested dict into dot-separated keys with string values."""
    items = {}
    for k, v in d.items():
        key = f"{prefix}{sep}{k}" if prefix else k
        if isinstance(v, dict):
            items.update(_flatten_dict(v, key, sep))
        elif isinstance(v, list):
            if v and isinstance(v[0], dict):
                # List of dicts — take first element
                items.update(_flatten_dict(v[0], key, sep))
            else:
                items[key] = str(v)
        else:
            items[key] = str(v) if v is not None else ""
    return items


# ---------------------------------------------------------------------------
# SQLite (stdlib)
# ---------------------------------------------------------------------------

def _load_sqlite(
    path: str,
    max_records: int,
    table: Optional[str],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load from a SQLite database file."""
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row

    try:
        if table is None:
            # Pick the table with the most rows
            cursor = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            )
            tables = [row[0] for row in cursor.fetchall()]
            if not tables:
                return [], []

            best_table = tables[0]
            best_count = 0
            for t in tables:
                try:
                    count = conn.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                    if count > best_count:
                        best_count = count
                        best_table = t
                except Exception:
                    continue
            table = best_table

        # Get field names
        cursor = conn.execute(f'PRAGMA table_info("{table}")')
        field_names = [row[1] for row in cursor.fetchall()]

        # Load records
        cursor = conn.execute(f'SELECT * FROM "{table}" LIMIT ?', (max_records,))
        records = []
        for row in cursor:
            record = {field_names[i]: str(row[i]) if row[i] is not None else ""
                      for i in range(len(field_names))}
            records.append(record)

        return records, field_names

    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Excel .xlsx (needs openpyxl)
# ---------------------------------------------------------------------------

def _load_excel(
    path: str,
    max_records: int,
    sheet: Optional[str],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load from an Excel .xlsx file. Requires openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "Excel support requires openpyxl: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    try:
        if sheet:
            ws = wb[sheet]
        else:
            ws = wb.active or wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)

        # First row is headers
        header_row = next(rows, None)
        if header_row is None:
            return [], []

        field_names = [str(h).strip() if h is not None else f"col_{i}"
                       for i, h in enumerate(header_row)]

        records = []
        for row in rows:
            if len(records) >= max_records:
                break
            record = {}
            for i, val in enumerate(row):
                if i < len(field_names):
                    record[field_names[i]] = str(val).strip() if val is not None else ""
            records.append(record)

        return records, field_names

    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Parquet (needs pyarrow)
# ---------------------------------------------------------------------------

def _load_parquet(
    path: str,
    max_records: int,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """Load from a Parquet file. Requires pyarrow."""
    try:
        import pyarrow.parquet as pq
    except ImportError:
        try:
            import pandas as pd
            df = pd.read_parquet(path)
            if len(df) > max_records:
                df = df.head(max_records)
            field_names = list(df.columns)
            records = []
            for _, row in df.iterrows():
                records.append({col: str(row[col]) if row[col] is not None else ""
                                for col in field_names})
            return records, field_names
        except ImportError:
            raise ImportError(
                "Parquet support requires pyarrow or pandas: "
                "pip install pyarrow  OR  pip install pandas"
            )

    table = pq.read_table(path)
    if table.num_rows > max_records:
        table = table.slice(0, max_records)

    field_names = table.column_names
    records = []
    for i in range(table.num_rows):
        record = {}
        for col in field_names:
            val = table.column(col)[i].as_py()
            record[col] = str(val) if val is not None else ""
        records.append(record)

    return records, field_names


# ---------------------------------------------------------------------------
# SQL dump parser (stdlib)
# ---------------------------------------------------------------------------

def _load_sql_dump(
    path: str,
    max_records: int,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse INSERT statements from a SQL dump file.
    Handles: INSERT INTO table (col1, col2) VALUES (val1, val2);
    """
    records = []
    field_names = []
    current_table = None

    # Regex for INSERT INTO table (columns) VALUES (values);
    insert_re = re.compile(
        r"INSERT\s+INTO\s+[`\"']?(\w+)[`\"']?\s*"
        r"\(([^)]+)\)\s*VALUES\s*"
        r"\((.+?)\)\s*;",
        re.IGNORECASE,
    )

    with open(path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            if len(records) >= max_records:
                break

            line = line.strip()
            if not line or line.startswith("--") or line.startswith("/*"):
                continue

            match = insert_re.match(line)
            if match:
                table_name = match.group(1)
                columns = [c.strip().strip("`\"'") for c in match.group(2).split(",")]
                values_str = match.group(3)

                # Parse values (handle quoted strings with commas)
                values = _parse_sql_values(values_str)

                if not field_names:
                    field_names = columns
                    current_table = table_name

                # Only load from the first table encountered
                if table_name == current_table and len(values) == len(columns):
                    record = {columns[i]: values[i] for i in range(len(columns))}
                    records.append(record)

    return records, field_names


def _parse_sql_values(values_str: str) -> List[str]:
    """Parse SQL VALUES clause, handling quoted strings."""
    values = []
    current = ""
    in_quote = False
    quote_char = ""

    for char in values_str:
        if in_quote:
            if char == quote_char:
                in_quote = False
            else:
                current += char
        elif char in ("'", '"'):
            in_quote = True
            quote_char = char
        elif char == ",":
            val = current.strip()
            if val.upper() == "NULL":
                val = ""
            values.append(val)
            current = ""
        else:
            current += char

    # Last value
    val = current.strip()
    if val.upper() == "NULL":
        val = ""
    values.append(val)

    return values


# ---------------------------------------------------------------------------
# Auto-detection utilities
# ---------------------------------------------------------------------------

def auto_detect_identity(
    records: List[Dict[str, Any]],
    field_names: List[str],
    max_candidates: int = 3,
) -> List[str]:
    """
    Pick identity fields by cardinality heuristic.

    Good identity fields have moderate cardinality: many distinct values
    but not unique per record (those are provenance/IDs).
    """
    total = len(records)
    if total == 0:
        return field_names[:2]

    scored = []
    for field in field_names:
        values = set(r.get(field, "") for r in records)
        unique_count = len(values)
        if unique_count > total * 0.8:
            continue
        if unique_count < 3:
            continue
        scored.append((unique_count, field))

    scored.sort(reverse=True)
    chosen = [field for _, field in scored[:max_candidates]]
    return chosen if chosen else field_names[:2]


def auto_detect_provenance(
    records: List[Dict[str, Any]],
    field_names: List[str],
) -> List[str]:
    """
    Detect fields that are likely record IDs (near-unique values).
    These should be excluded from routing.
    """
    total = len(records)
    if total == 0:
        return []

    provenance = []
    for field in field_names:
        values = set(r.get(field, "") for r in records)
        if len(values) > total * 0.8:
            provenance.append(field)
    return provenance
